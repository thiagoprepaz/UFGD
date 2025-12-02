#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automatização para Macro de Pagamentos — GUI (.ods -> 2 arquivos separados)

>>> Alteração solicitada (somente salvamento, lógica intacta):
- Salva em DOIS arquivos:
  1) "Progressão por Mérito UFGD.xlsx"  (linhas 13..63)
  2) "Progressão por Mérito HU.xlsx"    (linhas 71..121)

Demais comportamentos de UI e lógica permanecem iguais:
- Processamento automático ao escolher o .ods.
- Banner com altura dinâmica (evita texto cortado).
- Instruções: "1. Selecione o arquivo .ods.  2. Clique em 'Abrir pasta de saída'."
- Status mostra apenas "Caminho: <arquivo1> | <arquivo2>".
- Regras de extração (inalteradas):
  * Só gera linha quando J/N/S tiverem valor > 0 após abs() e round(2) E quando A ou B tiver conteúdo.
  * Para cada J/N/S com valor: A_r, B_r, MES/ANO=C5, rubrica="00001", rendimento="r",
    sequência=C6, valor, justificativa=C9, documento legal=C10.
  * 'valor' com 2 casas, formato '#,##0.00' (sem "R$"), coluna 'valor' largura 14.
"""

import os, re, sys, platform
from typing import Optional, Any, List, Dict

# -------- Lógica: dependências --------
try:
    import pandas as pd
except ModuleNotFoundError:
    import tkinter as _tk
    from tkinter import messagebox as _mb
    _tk.Tk().withdraw()
    _mb.showerror("Dependência ausente", "Instale:\n\npip install pandas odfpy openpyxl")
    sys.exit(1)

# -------- GUI --------
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter import font as tkfont
from openpyxl.utils import get_column_letter

# ================== UTIL / LÓGICA (inalterada) ==================
def col_to_index(col: str) -> int:
    col = col.strip().upper(); idx = 0
    for ch in col:
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Coluna inválida: {col}")
        idx = idx*26 + (ord(ch)-ord('A')+1)
    return idx-1

def get_cell(df: pd.DataFrame, col_letter: str, row_number: int) -> Any:
    r = row_number-1; c = col_to_index(col_letter)
    try: return df.iat[r, c]
    except Exception: return None

def parse_br_number(v: Any) -> Optional[float]:
    if v is None: return None
    if isinstance(v, (int, float)):
        try: return float(v)
        except: return None
    s = str(v).strip()
    if s == "" or s.lower() in {"nan","none"}: return None
    s = s.replace("R$","").replace("r$","").replace(" ","").replace("\xa0","")
    s = s.replace("%","").replace(".","").replace(",",".")
    s = re.sub(r"[^0-9\.\-]","", s)
    try: return float(s)
    except: return None

def not_blank(x: Any) -> bool:
    return x is not None and str(x).strip() != ""

def to_amount(v: Any) -> Optional[float]:
    f = parse_br_number(v)
    if f is None: return None
    val = round(abs(float(f)), 2)
    return val if val > 0 else None

def append_rows_for(df: pd.DataFrame, out: List[Dict[str, Any]], rownum: int):
    a = get_cell(df,"A",rownum)
    b = get_cell(df,"B",rownum)
    mes_ano = get_cell(df,"C",5)
    seq = get_cell(df,"C",6)
    just = get_cell(df,"C",9)
    doc = get_cell(df,"C",10)

    vals = {"J": to_amount(get_cell(df,"J",rownum)),
            "N": to_amount(get_cell(df,"N",rownum)),
            "S": to_amount(get_cell(df,"S",rownum))}
    has_amount = any(v is not None for v in vals.values())
    has_identity = not_blank(a) or not_blank(b)
    if not (has_amount and has_identity):
        return
    for _, num in vals.items():
        if num is not None:
            out.append({
                "A13": a, "B13": b, "MES/ANO": mes_ano,
                "rubrica": "00001", "rendimento": "r",
                "sequência": seq, "valor": num,
                "justificativa": just, "documento legal": doc,
            })

def process_sheet(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """(Mantida) Retorna todas as linhas (13..63 e 71..121)."""
    linhas: List[Dict[str, Any]] = []
    for r in list(range(13, 64)) + list(range(71, 122)):
        append_rows_for(df, linhas, r)
    return linhas

def process_sheet_dual(df: pd.DataFrame):
    """NOVO: separa por faixas sem alterar a lógica de extração."""
    linhas_ufgd: List[Dict[str, Any]] = []
    linhas_hu:   List[Dict[str, Any]] = []
    for r in range(13, 64):
        append_rows_for(df, linhas_ufgd, r)
    for r in range(71, 122):
        append_rows_for(df, linhas_hu, r)
    return linhas_ufgd, linhas_hu

def build_tables_and_counts_dual(file_path: str):
    """Retorna (df_total, details, df_ufgd, df_hu) — counts continuam por planilha."""
    if not os.path.exists(file_path): raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
    try: xls = pd.ExcelFile(file_path, engine="odf")
    except Exception as e: raise RuntimeError("Falha ao abrir .ods. Instale: pip install odfpy\n\nDetalhe: "+str(e)) from e

    all_rows: List[Dict[str, Any]] = []
    details: List[tuple] = []
    all_ufgd: List[Dict[str, Any]] = []
    all_hu:   List[Dict[str, Any]] = []

    for name in xls.sheet_names:
        df = pd.read_excel(file_path, sheet_name=name, header=None, engine="odf")
        ufgd_rows, hu_rows = process_sheet_dual(df)
        # para a árvore: total por planilha
        details.append((name, len(ufgd_rows) + len(hu_rows)))
        # acumula
        all_rows.extend(ufgd_rows); all_rows.extend(hu_rows)
        all_ufgd.extend(ufgd_rows); all_hu.extend(hu_rows)

    cols = ["A13","B13","MES/ANO","rubrica","rendimento","sequência","valor","justificativa","documento legal"]
    df_total = pd.DataFrame(all_rows, columns=cols)
    df_ufgd  = pd.DataFrame(all_ufgd, columns=cols)
    df_hu    = pd.DataFrame(all_hu, columns=cols)
    return df_total, details, df_ufgd, df_hu

def salvar_excel_as(df: pd.DataFrame, origem: str, basename: str) -> str:
    """Salva XLSX com nome definido por 'basename'."""
    base = os.path.dirname(os.path.abspath(origem))
    xlsx_out = os.path.join(base, basename)
    with pd.ExcelWriter(xlsx_out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="resultado")
        ws = writer.sheets["resultado"]
        if "valor" in df.columns:
            cidx = df.columns.get_loc("valor") + 1
            col_letter = get_column_letter(cidx)
            for r in range(2, len(df)+2):
                cell = ws[f"{col_letter}{r}"]
                cell.number_format = "#,##0.00"   # 1.234,56 (sem símbolo)
                try: cell.style = "Normal"
                except Exception: pass
            ws.column_dimensions[col_letter].width = 14
        for i, name in enumerate(df.columns, start=1):
            if name == "valor": continue
            try:
                maxlen = max((len(str(x)) for x in [name] + df[name].astype(str).tolist()), default=10)
            except Exception:
                maxlen = 15
            ws.column_dimensions[get_column_letter(i)].width = min(max(10, maxlen+2), 60)
    return xlsx_out

def abrir_pasta(path: str):
    try:
        if platform.system() == "Windows":
            os.startfile(path)  # type: ignore[attr-defined]
        elif platform.system() == "Darwin":
            import subprocess; subprocess.Popen(["open", path])
        else:
            import subprocess; subprocess.Popen(["xdg-open", path])
    except Exception:
        pass

# ================== APP (VISUAL preservado) ==================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Automatização para Macro de Pagamentos")
        self.geometry("980x560")
        self.resizable(False, False)
        self.configure(bg="#f5f5f5")
        self.ods_path: Optional[str] = None
        self.logo_img = None

        # Estilos (destaques)
        self.style = ttk.Style(self)
        try: self.style.theme_use("clam")
        except Exception: pass
        self.style.configure("Primary.TButton", foreground="white", background="#2E7D32",
                             font=("Segoe UI", 9, "bold"), padding=6)
        self.style.map("Primary.TButton", background=[("active", "#27692B"), ("disabled", "#A5D3A8")])
        self.style.configure("Muted.TButton", foreground="#333", background="#E0E0E0",
                             font=("Segoe UI", 9), padding=6)
        self.style.map("Muted.TButton", background=[("active", "#D5D5D5")])

        # ---- Banner (altura dinâmica) ----
        banner = tk.Frame(self, bg="#eaf5e6"); banner.pack(fill="x", side="top"); banner.pack_propagate(False)
        left = tk.Frame(banner, bg="#eaf5e6"); left.pack(side="left", padx=18, pady=14, fill="both", expand=True)

        title_font = tkfont.Font(family="Segoe UI", size=20, weight="bold")
        subtitle_black = tkfont.Font(family="Segoe UI", size=12, weight="bold")
        subtitle_green = tkfont.Font(family="Segoe UI", size=12)  # DPP maior

        tk.Label(left, text="Automatização para Macro de Pagamentos",
                 bg="#eaf5e6", fg="#1b1b1b", font=title_font).pack(anchor="w")
        tk.Label(left, text="Sequência 6 - 00001",
                 bg="#eaf5e6", fg="#000000", font=subtitle_black).pack(anchor="w", pady=(3, 0))
        tk.Label(left, text="DPP - Divisão de Pagamento de Pessoal",
                 bg="#eaf5e6", fg="#2e7d32", font=subtitle_green).pack(anchor="w", pady=(2, 0))

        right = tk.Frame(banner, bg="#eaf5e6"); right.pack(side="right", padx=18, pady=10)
        self._load_and_place_logo(right, target_h=80)

        title_h = title_font.metrics("linespace")
        sub1_h  = subtitle_black.metrics("linespace")
        sub2_h  = subtitle_green.metrics("linespace")
        needed  = 14 + title_h + 3 + sub1_h + 2 + sub2_h + 14
        banner.configure(height=max(needed, 120))

        # ---- Instruções ----
        hint = tk.Frame(self, bg="#e9e9e9", height=32); hint.pack(fill="x", padx=16, pady=(10, 10))
        tk.Label(hint, text="1. Selecione o arquivo .ods.  2. Clique em 'Abrir pasta de saída'.",
                 bg="#e9e9e9", fg="#333333", font=("Segoe UI", 10)).pack(anchor="w", padx=10, pady=6)

        # ---- Ações ----
        actions = tk.Frame(self, bg="#f5f5f5"); actions.pack(fill="x", padx=16)
        self.btn_browse = ttk.Button(actions, text="Selecionar arquivo .ods…",
                                     command=self.escolher_arquivo, style="Primary.TButton")
        self.btn_browse.pack(side="left")
        self.btn_abrir_pasta = ttk.Button(actions, text="Abrir pasta de saída",
                                          command=self.abrir_saida, state="disabled", style="Muted.TButton")
        self.btn_abrir_pasta.pack(side="left", padx=(8, 0))

        # ---- Status ----
        self.lbl_status = tk.Label(self, text="Nenhum arquivo selecionado.", anchor="w",
                                   fg="#333", bg="#f5f5f5")
        self.lbl_status.pack(fill="x", padx=18, pady=(10, 6))

        # ---- Tabela ----
        table_frame = tk.Frame(self, bg="#f5f5f5")
        table_frame.pack(fill="both", expand=True, padx=18, pady=(0, 10))
        self.tree = ttk.Treeview(table_frame, columns=("arquivo", "planilha", "linhas"),
                                 show="headings", height=12)
        self.tree.heading("arquivo", text="Arquivo");   self.tree.column("arquivo", width=420, anchor="w")
        self.tree.heading("planilha", text="Planilha"); self.tree.column("planilha", width=300, anchor="w")
        self.tree.heading("linhas", text="Linhas geradas"); self.tree.column("linhas", width=140, anchor="center")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns")
        table_frame.grid_columnconfigure(0, weight=1); table_frame.grid_rowconfigure(0, weight=1)

        # ---- Rodapé ----
        footer = tk.Frame(self, bg="#f5f5f5"); footer.pack(fill="x", side="bottom", padx=16, pady=(0, 10))
        tk.Label(footer, text="Versão 1.0 - 11/11/2025", bg="#f5f5f5", fg="#777").pack(side="left")
        tk.Button(footer, text="bom dia de trabalho", relief="groove", state="disabled",
                  fg="#2e7d32", bg="#eef7ea", disabledforeground="#2e7d32").pack(side="right")

    # ----- Logo helper -----
    def _load_and_place_logo(self, parent: tk.Widget, target_h: int = 80):
        paths = [
            os.path.join(os.getcwd(), "progesp-logo.png"),
            os.path.join(os.path.dirname(os.path.abspath(sys.argv[0])), "progesp-logo.png"),
        ]
        p = next((pp for pp in paths if os.path.exists(pp)), None)
        if not p:
            tk.Label(parent, text="PROGESP", fg="#2e7d32", bg="#eaf5e6",
                     font=tkfont.Font(family="Segoe UI", size=16, weight="bold")).grid(row=0, column=0, sticky="e")
            return
        try:
            from PIL import Image, ImageTk  # type: ignore
            im = Image.open(p)
            ratio = target_h / im.height
            new_w = max(1, int(im.width * ratio)); new_h = max(1, int(im.height * ratio))
            im = im.resize((new_w, new_h), Image.LANCZOS)
            self.logo_img = ImageTk.PhotoImage(im)
        except Exception:
            try:
                img = tk.PhotoImage(file=p)
                factor = max(1, int(round(img.height() / float(target_h))))
                self.logo_img = img.subsample(factor, factor)
            except Exception:
                self.logo_img = None
        if self.logo_img is not None:
            tk.Label(parent, image=self.logo_img, bg="#eaf5e6").grid(row=0, column=0, sticky="e")
        else:
            tk.Label(parent, text="PROGESP", fg="#2e7d32", bg="#eaf5e6",
                     font=tkfont.Font(family="Segoe UI", size=16, weight="bold")).grid(row=0, column=0, sticky="e")

    # ======= Handlers (processamento automático) =======
    def escolher_arquivo(self):
        p = filedialog.askopenfilename(title="Selecione o arquivo .ods",
                                       filetypes=[("Planilhas ODS","*.ods"), ("Todos os arquivos","*.*")])
        if not p: return
        self.ods_path = p
        self.btn_browse.configure(style="Primary.TButton")
        self.btn_abrir_pasta.configure(style="Muted.TButton", state="disabled")
        self.lbl_status.config(text=f"Processando '{p}' ...")
        for item in self.tree.get_children(): self.tree.delete(item)
        self.after(50, self.processar_automatico)

    def processar_automatico(self):
        try:
            # ---> usa versão dual (sem alterar lógica de extração)
            df_total, details, df_ufgd, df_hu = build_tables_and_counts_dual(self.ods_path)

            # Salva dois arquivos separados com nomes solicitados
            xlsx1 = salvar_excel_as(df_ufgd, self.ods_path, "Progressão por Mérito UFGD.xlsx")
            xlsx2 = salvar_excel_as(df_hu,   self.ods_path, "Progressão por Mérito HU.xlsx")

            # Popula a tabela (continua por planilha)
            arquivo = os.path.basename(self.ods_path)
            for (sheet_name, count) in details:
                self.tree.insert("", "end", values=(arquivo, sheet_name, count))

            # Status: mostra ambos os caminhos
            self.lbl_status.config(text=f"Caminho: {xlsx1}  |  {xlsx2}")

            # Destaques de botões pós-processamento
            self.btn_browse.configure(style="Muted.TButton")
            self.btn_abrir_pasta.configure(style="Primary.TButton", state="normal")

        except Exception as e:
            messagebox.showerror("Erro ao processar",
                                 f"Ocorreu um erro:\n\n{e}\n\nInstale dependências:\n pip install pandas odfpy openpyxl")
            self.lbl_status.config(text=f"Erro: {e}")
            self.btn_browse.configure(style="Primary.TButton")
            self.btn_abrir_pasta.configure(style="Muted.TButton", state="disabled")

    def abrir_saida(self):
        if self.ods_path:
            abrir_pasta(os.path.dirname(os.path.abspath(self.ods_path)))

if __name__ == "__main__":
    app = App()
    app.mainloop()